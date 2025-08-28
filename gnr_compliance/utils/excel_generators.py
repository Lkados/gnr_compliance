"""
Générateurs Excel pour les déclarations GNR avec formats exacts
"""

import frappe
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from typing import List, Dict, Any
import io

class GNRExcelGenerator:
    """Classe de base pour générer les fichiers Excel GNR avec formatage exact"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
    def set_column_widths(self, widths: Dict[str, float]):
        """Définir les largeurs de colonnes exactes"""
        for col_letter, width in widths.items():
            self.ws.column_dimensions[col_letter].width = width
            
    def set_row_height(self, row_num: int, height: float):
        """Définir la hauteur de ligne exacte"""
        self.ws.row_dimensions[row_num].height = height
        
    def merge_and_set_cell(self, range_str: str, value: str, font_size: int = 11, 
                          bold: bool = False, alignment: str = "center"):
        """Fusionner les cellules et définir la valeur avec formatage"""
        self.ws.merge_cells(range_str)
        cell = self.ws[range_str.split(':')[0]]
        cell.value = value
        cell.font = Font(size=font_size, bold=bold)
        cell.alignment = Alignment(horizontal=alignment, vertical="center")
        
    def set_cell_style(self, cell_ref: str, value: Any = None, font_size: int = 10, 
                      bold: bool = False, alignment: str = "center", 
                      background_color: str = None, border: bool = False):
        """Définir le style d'une cellule individuelle"""
        cell = self.ws[cell_ref]
        if value is not None:
            cell.value = value
        cell.font = Font(size=font_size, bold=bold)
        cell.alignment = Alignment(horizontal=alignment, vertical="center", wrap_text=True)
        
        if background_color:
            cell.fill = PatternFill(start_color=background_color, end_color=background_color, fill_type="solid")
            
        if border:
            thin_border = Side(style='thin')
            cell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def save_to_bytes(self) -> bytes:
        """Sauvegarder le workbook en bytes pour téléchargement"""
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()


class ArreteTrimestrielGenerator(GNRExcelGenerator):
    """Générateur pour l'Arrêté Trimestriel de Stock avec format exact"""
    
    def generate(self, period_start: str, period_end: str, company_name: str, 
                autorisation_number: str, stock_movements: List[Dict]) -> bytes:
        """
        Générer l'arrêté trimestriel avec le format exact du fichier original
        
        Args:
            period_start: Date de début (YYYY-MM-DD)
            period_end: Date de fin (YYYY-MM-DD) 
            company_name: Nom de la société
            autorisation_number: Numéro d'autorisation
            stock_movements: Liste des mouvements de stock quotidiens
        """
        
        # Configuration des largeurs de colonnes exactes
        column_widths = {
            'A': 17.71428571429,
            'B': 11.71428571429, 
            'C': 11.71428571429,
            'D': 11.71428571429,
            'E': 11.71428571429,
            'F': 11.71428571429,
            'G': 11.71428571429
        }
        self.set_column_widths(column_widths)
        
        # En-têtes principaux (lignes 1-4)
        self.set_row_height(1, 24)
        self.merge_and_set_cell('A1:G1', 'Comptabilité Matière - Gasoil Non Routier', 
                               font_size=12, bold=True)
        
        self.set_row_height(2, 23.25)
        self.merge_and_set_cell('A2:G2', f'Société : {company_name}', 
                               font_size=11, bold=True)
        
        self.set_row_height(3, 19.5)
        self.merge_and_set_cell('A3:G3', f'Numéro d\'autorisation : {autorisation_number}', 
                               font_size=11, bold=True)
        
        # Déterminer le trimestre et l'année
        start_date = datetime.strptime(period_start, '%Y-%m-%d')
        end_date = datetime.strptime(period_end, '%Y-%m-%d')
        
        quarter_text = self._get_quarter_text(start_date, end_date)
        self.set_row_height(4, 19.5)
        self.merge_and_set_cell('A4:G4', quarter_text, font_size=11, bold=True)
        
        # Ligne vide
        self.set_row_height(5, 15.75)
        
        # Sous-titre "Volume réel en Litres"
        self.merge_and_set_cell('B6:G6', 'Volume réel en Litres', font_size=11, bold=True)
        
        # En-têtes des colonnes (lignes 7-8)
        # Ligne 7 - En-têtes principaux
        self.set_cell_style('A7', 'Date', bold=True, border=True)
        self.set_cell_style('B7', 'Stock Initial', bold=True, border=True)
        self.merge_and_set_cell('C7:D7', 'Entrées', font_size=10, bold=True)
        self.set_cell_style('E7', 'Sorties', bold=True, border=True)
        self.set_cell_style('G7', 'Stock Final', bold=True, border=True)
        
        # Ligne 8 - Sous-en-têtes
        self.set_row_height(8, 50.25)
        self.set_cell_style('C8', 'N° BL', bold=True, border=True)
        self.set_cell_style('D8', 'Volume', bold=True, border=True)
        self.set_cell_style('E8', 'AGRICOLE /\nFORESTIER\nVolume', bold=True, border=True)
        self.set_cell_style('F8', 'Sans\nAttestation\nVolume', bold=True, border=True)
        
        # Données des mouvements de stock (à partir de la ligne 9)
        current_row = 9
        running_stock = 0
        
        for movement in stock_movements:
            # Convertir la date en nombre Excel
            movement_date = datetime.strptime(movement['date'], '%Y-%m-%d')
            excel_date = self._date_to_excel(movement_date)
            
            # Date
            self.set_cell_style(f'A{current_row}', excel_date, alignment='center', border=True)
            
            # Stock initial (report du stock final précédent)
            if current_row == 9:
                running_stock = movement.get('stock_initial', 0)
            self.set_cell_style(f'B{current_row}', running_stock, alignment='center', border=True)
            
            # Entrées - N° BL (optionnel)
            bl_number = movement.get('bl_number', '')
            if bl_number:
                self.set_cell_style(f'C{current_row}', bl_number, alignment='center', border=True)
            
            # Entrées - Volume
            entrees = movement.get('entrees', 0)
            self.set_cell_style(f'D{current_row}', entrees, alignment='center', border=True)
            
            # Sorties - Agricole/Forestier
            sorties_agricole = movement.get('sorties_agricole', 0)
            self.set_cell_style(f'E{current_row}', sorties_agricole, alignment='center', border=True)
            
            # Sorties - Sans Attestation  
            sorties_sans_attestation = movement.get('sorties_sans_attestation', 0)
            self.set_cell_style(f'F{current_row}', sorties_sans_attestation, alignment='center', border=True)
            
            # Stock Final - Formule Excel
            stock_final_formula = f'=B{current_row}+D{current_row}-E{current_row}-F{current_row}'
            self.ws[f'G{current_row}'].value = stock_final_formula
            self.set_cell_style(f'G{current_row}', alignment='center', border=True)
            
            # Calculer le stock pour la ligne suivante
            running_stock = running_stock + entrees - sorties_agricole - sorties_sans_attestation
            
            current_row += 1
        
        # Ligne de cumuls (equivalent à la ligne 68 dans l'original)
        cumul_row = current_row + 1
        self.set_row_height(cumul_row, 15.75)
        self.set_cell_style(f'A{cumul_row}', 'Cumul Trimestriel', bold=True, border=True)
        
        # Formules de somme pour les cumuls
        start_data_row = 9
        end_data_row = current_row - 1
        
        self.ws[f'D{cumul_row}'].value = f'=SUM(D{start_data_row}:D{end_data_row})'
        self.ws[f'E{cumul_row}'].value = f'=SUM(E{start_data_row}:E{end_data_row})'
        self.ws[f'F{cumul_row}'].value = f'=SUM(F{start_data_row}:F{end_data_row})'
        
        self.set_cell_style(f'D{cumul_row}', border=True, bold=True)
        self.set_cell_style(f'E{cumul_row}', border=True, bold=True)
        self.set_cell_style(f'F{cumul_row}', border=True, bold=True)
        
        # Récapitulatif final (lignes 70-72)
        recap_start = cumul_row + 2
        
        # Stock comptable
        self.merge_and_set_cell(f'D{recap_start}:F{recap_start}', 'Stock comptable au 31 Mars 2025', 
                               bold=True, alignment='center')
        self.ws[f'G{recap_start}'].value = f'=G{end_data_row}'
        self.set_cell_style(f'G{recap_start}', border=True, bold=True)
        
        # Stock physique 
        self.merge_and_set_cell(f'D{recap_start+1}:F{recap_start+1}', 'Stock physique au 31 Mars 2025', 
                               bold=True, alignment='center')
        # Valeur à remplir manuellement ou depuis les données
        final_physical_stock = stock_movements[-1].get('stock_physique', running_stock) if stock_movements else 0
        self.set_cell_style(f'G{recap_start+1}', final_physical_stock, border=True, bold=True)
        
        # Écart
        self.set_row_height(recap_start+2, 15.75)
        self.merge_and_set_cell(f'D{recap_start+2}:F{recap_start+2}', 'Écart', 
                               bold=True, alignment='center')
        self.ws[f'G{recap_start+2}'].value = f'=G{recap_start+1}-G{recap_start}'
        self.set_cell_style(f'G{recap_start+2}', border=True, bold=True)
        
        return self.save_to_bytes()
    
    def _get_quarter_text(self, start_date: datetime, end_date: datetime) -> str:
        """Générer le texte du trimestre"""
        quarters = {
            1: "1er Trimestre {} (Janvier - Février - Mars)",
            2: "2ème Trimestre {} (Avril - Mai - Juin)", 
            3: "3ème Trimestre {} (Juillet - Août - Septembre)",
            4: "4ème Trimestre {} (Octobre - Novembre - Décembre)"
        }
        quarter = (start_date.month - 1) // 3 + 1
        return quarters[quarter].format(start_date.year)
    
    def _date_to_excel(self, date: datetime) -> int:
        """Convertir une date Python en numéro de série Excel"""
        excel_epoch = datetime(1899, 12, 30)
        delta = date - excel_epoch
        return delta.days


class ListeClientsGenerator(GNRExcelGenerator):
    """Générateur pour la Liste Semestrielle des Clients avec format exact"""
    
    def generate(self, period_start: str, period_end: str, company_name: str,
                company_siren: str, clients_data: List[Dict]) -> bytes:
        """
        Générer la liste semestrielle des clients avec le format exact
        
        Args:
            period_start: Date de début (YYYY-MM-DD)
            period_end: Date de fin (YYYY-MM-DD)
            company_name: Nom de la société distributeur
            company_siren: SIREN du distributeur
            clients_data: Liste des données clients
        """
        
        # Configuration des largeurs de colonnes exactes
        column_widths = {
            'A': 32.71428571429,  # Raison sociale distributeur
            'B': 18.71428571429,  # SIREN distributeur
            'C': 47.57142857143,  # Raison sociale client
            'D': 18.71428571429,  # SIREN client
            'E': 34.71428571429,  # Volumes
            'F': 34.71428571429   # Tarif d'accise
        }
        self.set_column_widths(column_widths)
        
        # Étendre à 400+ colonnes comme dans l'original (template très large)
        for col_num in range(7, 401):
            col_letter = get_column_letter(col_num)
            self.ws.column_dimensions[col_letter].width = 11.35714285714
        
        # En-têtes principaux (ligne 1)
        self.set_cell_style('A1', 'Informations distributeur autorisé ou fournisseur', 
                           font_size=11, bold=True, alignment='center')
        self.set_cell_style('C1', 'Informations client', 
                           font_size=11, bold=True, alignment='center')
        self.set_cell_style('E1', 'Données carburant GNR', 
                           font_size=11, bold=True, alignment='center')
        
        # En-têtes des colonnes (ligne 2)
        self.set_cell_style('A2', 'Raison sociale', font_size=10, bold=True, border=True)
        self.set_cell_style('B2', 'SIREN', font_size=10, bold=True, border=True)
        self.set_cell_style('C2', 'Raison sociale du client', font_size=10, bold=True, border=True)
        self.set_cell_style('D2', 'SIREN du client', font_size=10, bold=True, border=True)
        self.set_cell_style('E2', 'Volumes (en hL) de GNR livrés au client au cours du dernier semestre civil écoulé', 
                           font_size=10, bold=True, border=True)
        self.set_cell_style('F2', 'Tarif d\'accise appliqué (en euro par hectolitres)', 
                           font_size=10, bold=True, border=True)
        
        # Données clients (à partir de la ligne 3)
        current_row = 3
        
        for client in clients_data:
            # Distributeur (toujours le même)
            self.set_cell_style(f'A{current_row}', company_name, border=True, alignment='left')
            
            # SIREN distributeur (si fourni)
            if company_siren:
                self.set_cell_style(f'B{current_row}', company_siren, border=True, alignment='center')
            
            # Raison sociale client
            self.set_cell_style(f'C{current_row}', client.get('raison_sociale', ''), 
                               border=True, alignment='left')
            
            # SIREN client
            client_siren = client.get('siren', '')
            if client_siren:
                self.set_cell_style(f'D{current_row}', client_siren, border=True, alignment='center')
            
            # Volume en hectolitres
            volume_hl = client.get('volume_hl', 0)
            self.set_cell_style(f'E{current_row}', volume_hl, border=True, alignment='center')
            
            # Tarif d'accise
            tarif_accise = client.get('tarif_accise', 0)
            self.set_cell_style(f'F{current_row}', tarif_accise, border=True, alignment='center')
            
            current_row += 1
        
        return self.save_to_bytes()


def generate_arrete_trimestriel(period_start: str, period_end: str) -> bytes:
    """
    API fonction pour générer l'arrêté trimestriel
    """
    # Récupérer les informations de la société
    company = frappe.defaults.get_user_default("Company") or frappe.get_all("Company", limit=1)[0].name
    company_doc = frappe.get_doc("Company", company)
    
    # Récupérer le numéro d'autorisation depuis les paramètres
    autorisation_number = frappe.db.get_single_value('GNR Settings', 'autorisation_number') or "08/2024/AMIENS"
    
    # Récupérer les mouvements de stock pour la période
    stock_movements = get_stock_movements_for_period(period_start, period_end)
    
    generator = ArreteTrimestrielGenerator()
    return generator.generate(
        period_start=period_start,
        period_end=period_end,
        company_name=company_doc.company_name,
        autorisation_number=autorisation_number,
        stock_movements=stock_movements
    )


def generate_liste_clients(period_start: str, period_end: str) -> bytes:
    """
    API fonction pour générer la liste semestrielle des clients
    """
    # Récupérer les informations de la société
    company = frappe.defaults.get_user_default("Company") or frappe.get_all("Company", limit=1)[0].name
    company_doc = frappe.get_doc("Company", company)
    
    # Récupérer les données clients pour la période
    clients_data = get_clients_data_for_period(period_start, period_end)
    
    generator = ListeClientsGenerator()
    return generator.generate(
        period_start=period_start,
        period_end=period_end,
        company_name=company_doc.company_name,
        company_siren=company_doc.tax_id or "",
        clients_data=clients_data
    )


def get_stock_movements_for_period(period_start: str, period_end: str) -> List[Dict]:
    """
    Récupérer les mouvements de stock pour la période donnée
    """
    movements = frappe.db.sql("""
        SELECT 
            date,
            SUM(CASE WHEN purpose = 'Receipt' THEN qty_in_liters ELSE 0 END) as entrees,
            SUM(CASE WHEN purpose = 'Material Issue' AND customer_category = 'Agricole' 
                THEN qty_in_liters ELSE 0 END) as sorties_agricole,
            SUM(CASE WHEN purpose = 'Material Issue' AND (customer_category != 'Agricole' OR customer_category IS NULL) 
                THEN qty_in_liters ELSE 0 END) as sorties_sans_attestation,
            bl_number,
            stock_physique
        FROM `tabMouvement GNR`
        WHERE date BETWEEN %s AND %s
        GROUP BY date
        ORDER BY date
    """, (period_start, period_end), as_dict=True)
    
    return movements or []


def get_clients_data_for_period(period_start: str, period_end: str) -> List[Dict]:
    """
    Récupérer les données clients pour la période donnée
    """
    clients = frappe.db.sql("""
        SELECT 
            c.customer_name as raison_sociale,
            c.tax_id as siren,
            SUM(m.qty_in_liters / 100) as volume_hl,
            CASE 
                WHEN m.customer_category = 'Agricole' THEN 3.86
                ELSE 24.81 
            END as tarif_accise
        FROM `tabMouvement GNR` m
        JOIN `tabCustomer` c ON m.customer = c.name
        WHERE m.date BETWEEN %s AND %s
        AND m.purpose = 'Material Issue'
        GROUP BY c.name, m.customer_category
        ORDER BY c.customer_name
    """, (period_start, period_end), as_dict=True)
    
    return clients or []