# gnr_compliance/utils/export_reglementaire.py
# Générateurs d'exports aux formats réglementaires français

import frappe
from frappe.utils import getdate, flt
from frappe.utils.xlutils import make_ods
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

@frappe.whitelist()
def generer_arrete_trimestriel(from_date, to_date, include_details=True):
    """
    Génère l'Arrêté Trimestriel de Stock Détaillé au format réglementaire
    
    Args:
        from_date: Date début (YYYY-MM-DD)
        to_date: Date fin (YYYY-MM-DD)
        include_details: Inclure les détails par produit
    
    Returns:
        dict: URL du fichier généré
    """
    try:
        # Récupérer les données des mouvements GNR
        mouvements = get_mouvements_gnr_periode(from_date, to_date)
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Arrêté Stock Détaillé"
        
        # === STYLES ===
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === EN-TÊTE ===
        periode_text = f"{datetime.strptime(from_date, '%Y-%m-%d').strftime('%B %Y')} à {datetime.strptime(to_date, '%Y-%m-%d').strftime('%B %Y')}"
        
        # Fusionner et écrire le titre
        ws.merge_cells('A1:H1')
        ws['A1'] = f'ARRÊTÉ TRIMESTRIEL DE STOCK DÉTAILLÉ - {periode_text}'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = header_alignment
        
        ws.merge_cells('A2:H2')
        ws['A2'] = 'Conformément à l\'arrêté du 28 juin 2001'
        ws['A2'].font = header_font
        ws['A2'].fill = header_fill
        ws['A2'].alignment = header_alignment
        
        # === EN-TÊTES COLONNES ===
        headers = [
            'Code Produit', 'Désignation', 'Stock Début (hL)',
            'Entrées (hL)', 'Sorties (hL)', 'Stock Fin (hL)',
            'Taux GNR (€/hL)', 'Montant Taxe (€)'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # === DONNÉES ===
        row = 5
        total_taxe = 0
        
        # Grouper par produit
        produits_data = {}
        for mouvement in mouvements:
            code = mouvement.code_produit
            if code not in produits_data:
                produits_data[code] = {
                    'designation': frappe.get_value('Item', code, 'item_name') or code,
                    'stock_debut': 0,
                    'entrees': 0,
                    'sorties': 0,
                    'taux_gnr': mouvement.taux_gnr or 0,
                    'montant_taxe': 0
                }
            
            if mouvement.type_mouvement in ['Achat', 'Entrée']:
                produits_data[code]['entrees'] += flt(mouvement.quantite)
            elif mouvement.type_mouvement in ['Vente', 'Sortie']:
                produits_data[code]['sorties'] += flt(mouvement.quantite)
            
            produits_data[code]['montant_taxe'] += flt(mouvement.montant_taxe_gnr or 0)
        
        # Écrire les données
        for code_produit, data in produits_data.items():
            data['stock_fin'] = data['stock_debut'] + data['entrees'] - data['sorties']
            
            ws.cell(row=row, column=1, value=code_produit).border = border
            ws.cell(row=row, column=2, value=data['designation']).border = border
            ws.cell(row=row, column=3, value=data['stock_debut']).border = border
            ws.cell(row=row, column=4, value=data['entrees']).border = border
            ws.cell(row=row, column=5, value=data['sorties']).border = border
            ws.cell(row=row, column=6, value=data['stock_fin']).border = border
            ws.cell(row=row, column=7, value=data['taux_gnr']).border = border
            ws.cell(row=row, column=8, value=data['montant_taxe']).border = border
            
            total_taxe += data['montant_taxe']
            row += 1
        
        # Total final
        ws.merge_cells(f'A{row+2}:G{row+2}')
        total_cell = ws[f'A{row+2}']
        total_cell.value = 'TOTAL TAXE GNR'
        total_cell.font = header_font
        total_cell.fill = header_fill
        total_cell.alignment = header_alignment
        
        ws.cell(row=row+2, column=8, value=total_taxe).border = border
        
        # Ajuster les largeurs de colonnes
        column_widths = [15, 30, 15, 15, 15, 15, 15, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Sauvegarder dans un buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Créer le fichier
        periode_nom = f"{datetime.strptime(from_date, '%Y-%m-%d').strftime('%Y %B')} à {datetime.strptime(to_date, '%Y-%m-%d').strftime('%B')}"
        file_name = f"TIPAccEne Arrêté Trimestriel de Stock Détaillé {periode_nom}.xlsx"
        
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": file_name,
            "content": output.getvalue()
        })
        file_doc.save()
        
        return {"file_url": file_doc.file_url, "file_name": file_name}
        
    except Exception as e:
        frappe.throw(f"Erreur génération arrêté trimestriel: {str(e)}")

@frappe.whitelist()
def generer_liste_semestrielle_clients(from_date, to_date):
    """
    Génère la Liste Semestrielle des Clients pour la Douane
    
    Args:
        from_date: Date début semestre
        to_date: Date fin semestre
    
    Returns:
        dict: URL du fichier généré
    """
    try:
        # Récupérer les données clients avec ventes GNR
        clients_data = get_clients_gnr_periode(from_date, to_date)
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Liste Clients Semestrielle"
        
        # === STYLES ===
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === EN-TÊTE ===
        periode_text = f"{datetime.strptime(from_date, '%Y-%m-%d').strftime('%B %Y')} à {datetime.strptime(to_date, '%Y-%m-%d').strftime('%B %Y')}"
        
        # Fusionner et écrire le titre
        ws.merge_cells('A1:F1')
        ws['A1'] = f'LISTE SEMESTRIELLE DES CLIENTS - {periode_text}'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = header_alignment
        
        ws.merge_cells('A2:F2')
        ws['A2'] = 'Déclaration pour la Direction Générale des Douanes et Droits Indirects'
        ws['A2'].font = header_font
        ws['A2'].fill = header_fill
        ws['A2'].alignment = header_alignment
        
        # === EN-TÊTES COLONNES ===
        headers = [
            'Code Client', 'Nom/Raison Sociale', 'SIRET',
            'Quantité Totale (hL)', 'Montant HT (€)', 'Montant Taxe GNR (€)'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # === DONNÉES CLIENTS ===
        row = 5
        total_quantite = 0
        total_ht = 0
        total_taxe = 0
        
        for client in clients_data:
            ws.cell(row=row, column=1, value=client['code_client']).border = border
            ws.cell(row=row, column=2, value=client['nom_client']).border = border
            ws.cell(row=row, column=3, value=client['siret'] or '').border = border
            ws.cell(row=row, column=4, value=client['quantite_totale']).border = border
            ws.cell(row=row, column=5, value=client['montant_ht']).border = border
            ws.cell(row=row, column=6, value=client['montant_taxe']).border = border
            
            total_quantite += client['quantite_totale']
            total_ht += client['montant_ht']
            total_taxe += client['montant_taxe']
            row += 1
        
        # Totaux
        ws.merge_cells(f'A{row+2}:C{row+2}')
        total_cell = ws[f'A{row+2}']
        total_cell.value = 'TOTAUX'
        total_cell.font = header_font
        total_cell.fill = header_fill
        total_cell.alignment = header_alignment
        
        ws.cell(row=row+2, column=4, value=total_quantite).border = border
        ws.cell(row=row+2, column=5, value=total_ht).border = border
        ws.cell(row=row+2, column=6, value=total_taxe).border = border
        
        # Ajuster les largeurs de colonnes
        column_widths = [15, 35, 20, 18, 18, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Informations légales
        ws.cell(row=row+5, column=1, value='Déclaration établie conformément au décret n°2001-387 du 3 mai 2001')
        ws.cell(row=row+6, column=1, value=f'Date d\'établissement: {datetime.now().strftime("%d/%m/%Y")}')
        
        # Sauvegarder dans un buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Créer le fichier
        periode_nom = f"{datetime.strptime(from_date, '%Y-%m-%d').strftime('%Y %B')} à {datetime.strptime(to_date, '%Y-%m-%d').strftime('%B')}"
        file_name = f"TIPAccEne Liste Semestrielle des Clients Douane {periode_nom}.xlsx"
        
        file_doc = frappe.get_doc({
            "doctype": "File", 
            "file_name": file_name,
            "content": output.getvalue()
        })
        file_doc.save()
        
        return {"file_url": file_doc.file_url, "file_name": file_name}
        
    except Exception as e:
        frappe.throw(f"Erreur génération liste clients: {str(e)}")

def get_mouvements_gnr_periode(from_date, to_date):
    """Récupère les mouvements GNR pour une période"""
    return frappe.db.sql("""
        SELECT 
            code_produit,
            type_mouvement,
            quantite,
            taux_gnr,
            montant_taxe_gnr
        FROM `tabMouvement GNR`
        WHERE date_mouvement BETWEEN %s AND %s
        AND docstatus = 1
        ORDER BY code_produit, date_mouvement
    """, (from_date, to_date), as_dict=True)

def get_clients_gnr_periode(from_date, to_date):
    """Récupère les données clients avec ventes GNR pour une période"""
    return frappe.db.sql("""
        SELECT 
            m.client as code_client,
            c.customer_name as nom_client,
            c.siret,
            SUM(m.quantite) as quantite_totale,
            SUM(m.quantite * m.prix_unitaire) as montant_ht,
            SUM(m.montant_taxe_gnr) as montant_taxe
        FROM `tabMouvement GNR` m
        LEFT JOIN `tabCustomer` c ON m.client = c.name
        WHERE m.date_mouvement BETWEEN %s AND %s
        AND m.docstatus = 1
        AND m.type_mouvement = 'Vente'
        AND m.client IS NOT NULL
        GROUP BY m.client
        HAVING SUM(m.quantite) > 0
        ORDER BY SUM(m.quantite) DESC
    """, (from_date, to_date), as_dict=True)